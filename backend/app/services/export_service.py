"""Service d'export Excel (openpyxl). Retourne les fichiers en bytes."""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

JAUNE = "FFF2B705"
ENTETE_FONT = Font(bold=True, color="FF1A1A1A")
ENTETE_FILL = PatternFill(start_color=JAUNE, end_color=JAUNE, fill_type="solid")


def _styliser_entete(ws, nb_colonnes: int):
    for col in range(1, nb_colonnes + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = ENTETE_FONT
        cell.fill = ENTETE_FILL
        cell.alignment = Alignment(horizontal="center")


def exporter_paiements(paiements: list[dict]) -> bytes:
    """paiements : liste de dicts (reference, date, client, type, montant, moyen, statut)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements"

    entetes = ["Référence", "Date", "Client", "Type", "Montant (MRU)", "Moyen", "Statut"]
    ws.append(entetes)
    _styliser_entete(ws, len(entetes))

    for p in paiements:
        ws.append([
            p.get("reference", ""),
            str(p.get("date_paiement", "")),
            p.get("client_nom", ""),
            p.get("type_paiement", ""),
            float(p.get("montant", 0)),
            p.get("moyen_paiement", ""),
            p.get("statut", ""),
        ])

    # Largeurs de colonnes
    for col, larg in zip("ABCDEFG", [22, 20, 25, 20, 15, 15, 12]):
        ws.column_dimensions[col].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def exporter_clients(clients: list[dict]) -> bytes:
    """clients : liste de dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    entetes = ["N° Membre", "Nom", "Prénom", "Sexe", "Téléphone", "Email", "Actif"]
    ws.append(entetes)
    _styliser_entete(ws, len(entetes))

    for c in clients:
        ws.append([
            c.get("numero_membre", ""),
            c.get("nom", ""),
            c.get("prenom", ""),
            c.get("sexe", ""),
            c.get("telephone", ""),
            c.get("email", ""),
            "Oui" if c.get("actif") else "Non",
        ])

    for col, larg in zip("ABCDEFG", [16, 18, 18, 10, 16, 25, 8]):
        ws.column_dimensions[col].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def exporter_rapport_detaille(feuille: str, entetes: list[str], lignes: list[list]) -> bytes:
    """Export générique vers Excel pour les rapports détaillés."""
    wb = Workbook()
    ws = wb.active
    ws.title = feuille[:31] if feuille else "Rapport"

    ws.append(entetes)
    _styliser_entete(ws, len(entetes))

    for ligne in lignes:
        ws.append(ligne)

    for idx, _ in enumerate(entetes, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generer_modele_import_clients() -> bytes:
    """Génère un fichier Excel modèle pour l'import en masse de clients."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import clients"

    entetes = [
        "Nom *",
        "Prénom *",
        "Téléphone *",
        "Sexe",
        "Date naissance",
        "WhatsApp",
        "Adresse",
        "Email",
        "N° pièce identité",
        "Contact urgence",
        "Tél. urgence",
        "Groupe sanguin",
    ]
    ws.append(entetes)
    _styliser_entete(ws, len(entetes))

    ws.append([
        "Diallo",
        "Amadou",
        "22246123456",
        "Homme",
        "15/03/1995",
        "22246123456",
        "Nouakchott",
        "amadou@email.com",
        "N123456",
        "Fatou Diallo",
        "22246987654",
        "O+",
    ])

    note = ws.cell(row=3, column=1, value="* Champs obligatoires. Sexe : Homme ou Femme. Date : JJ/MM/AAAA.")
    note.font = Font(italic=True, color="FF64748B")

    for col, larg in zip("ABCDEFGHIJKL", [16, 16, 16, 10, 16, 16, 22, 24, 16, 18, 14, 14]):
        ws.column_dimensions[col].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generer_modele_import_abonnements() -> bytes:
    """Génère un modèle Excel pour import/saisie d'abonnements."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import abonnements"

    entetes = [
        "N° membre client *",
        "Moyen paiement (code) *",
        "Note",
    ]
    ws.append(entetes)
    _styliser_entete(ws, len(entetes))

    ws.append(["TF-2026-0001", "cash", "Souscription ou renouvellement"])

    note = ws.cell(
        row=3,
        column=1,
        value="* Champs obligatoires. Codes moyens de paiement: cash, bankily, masrivi, sedad, amanty, click, bimbank, bcipay, moov_money.",
    )
    note.font = Font(italic=True, color="FF64748B")

    for col, larg in zip("ABC", [22, 24, 44]):
        ws.column_dimensions[col].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generer_modele_import_paiements() -> bytes:
    """Génère un modèle Excel pour import/saisie de paiements."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import paiements"

    entetes = [
        "Type paiement *",
        "N° membre client",
        "Moyen paiement (code) *",
        "Montant (MRU)",
        "Nom client occasionnel",
    ]
    ws.append(entetes)
    _styliser_entete(ws, len(entetes))

    ws.append(["Abonnement", "TF-2026-0001", "cash", "", ""])
    ws.append(["Séance journalière", "", "cash", "50", "Client occasionnel"])

    note = ws.cell(
        row=4,
        column=1,
        value="* Type paiement obligatoire: Abonnement, Séance journalière, Service supplémentaire, Autre.",
    )
    note.font = Font(italic=True, color="FF64748B")

    for col, larg in zip("ABCDE", [22, 22, 24, 16, 26]):
        ws.column_dimensions[col].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
