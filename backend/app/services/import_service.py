"""Import Excel des clients."""
import io
import uuid
from datetime import date, datetime

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.models.client import Client
from app.schemas.client import ClientCreate
from app.services.client_service import generer_numero_membre


def _cell_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Date invalide « {text} » (format attendu : JJ/MM/AAAA).")


def _parse_row(row: tuple, ligne: int) -> ClientCreate:
    cells = list(row) + [None] * (12 - len(row))

    nom = _cell_str(cells[0])
    prenom = _cell_str(cells[1])
    telephone = _cell_str(cells[2])

    if not nom:
        raise ValueError("Le nom est obligatoire.")
    if not prenom:
        raise ValueError("Le prénom est obligatoire.")
    if not telephone:
        raise ValueError("Le téléphone est obligatoire.")

    sexe = _cell_str(cells[3])
    if sexe:
        sexe_norm = sexe.capitalize()
        if sexe_norm not in ("Homme", "Femme"):
            raise ValueError("Le sexe doit être « Homme » ou « Femme ».")
        sexe = sexe_norm

    email = _cell_str(cells[7])
    if email in (None, "-", "—"):
        email = None

    return ClientCreate(
        nom=nom,
        prenom=prenom,
        telephone=telephone,
        sexe=sexe,
        date_naissance=_parse_date(cells[4]),
        whatsapp=_cell_str(cells[5]),
        adresse=_cell_str(cells[6]),
        email=email,
        numero_piece_identite=_cell_str(cells[8]),
        contact_urgence_nom=_cell_str(cells[9]),
        contact_urgence_telephone=_cell_str(cells[10]),
        groupe_sanguin=_cell_str(cells[11]),
    )


def importer_excel(
    db: Session,
    file_bytes: bytes,
    created_by: uuid.UUID,
) -> dict:
    """Importe des clients depuis un fichier Excel (.xlsx)."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Fichier Excel invalide ou illisible.") from exc

    ws = wb.active
    if ws is None:
        raise ValueError("Le fichier Excel ne contient aucune feuille.")

    erreurs: list[dict] = []
    crees = 0
    total = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        first_cell = _cell_str(row[0] if row else None) or ""
        if first_cell.startswith("*") or first_cell.lower().startswith("champs obligatoires"):
            continue

        total += 1
        try:
            data = _parse_row(row, idx)
            client = Client(
                numero_membre=generer_numero_membre(db),
                created_by=created_by,
                **data.model_dump(),
            )
            db.add(client)
            db.flush()
            crees += 1
        except ValidationError as exc:
            msg = exc.errors()[0]["msg"] if exc.errors() else str(exc)
            erreurs.append({"ligne": idx, "message": msg})
        except ValueError as exc:
            erreurs.append({"ligne": idx, "message": str(exc)})

    wb.close()

    if crees > 0:
        db.commit()
    else:
        db.rollback()

    return {
        "total_lignes": total,
        "crees": crees,
        "echecs": len(erreurs),
        "erreurs": erreurs,
    }
